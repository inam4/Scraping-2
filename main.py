from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup
import time
import re

from selenium.webdriver.support.ui import Select
from openpyxl import Workbook


wb = Workbook()

driver = webdriver.Chrome(executable_path=r"C:\Users\Hp\Desktop\Coursera\Driver\chromedriver.exe")



driver.get("https://suissetec.ch/de/gebaeudetechniker_finden.html")
actions = ActionChains(driver)



plzs = ['1000','2000','3000','4000','5000','6000','7000','8000','9000']
time.sleep(10)
count=0
for i in plzs:
    count=count+1
    sh1=wb.create_sheet(i,count)
    sh1.cell(row=1, column=1).value = "Name"
    sh1.cell(row=1, column=2).value = "PLZ"
    sh1.cell(row=1, column=3).value = "Address"
    sh1.cell(row=1, column=4).value = "Phone"
    sh1.cell(row=1, column=5).value = "Email"
    sh1.cell(row=1, column=6).value = "Website"
    e2 = driver.find_element_by_xpath("//*[@id='filter_plz']")
    e3 = driver.find_element_by_xpath('//*[@id="filter_gmap"]/button')
    e2.send_keys(i)
    e3.click()
    time.sleep(30)
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    e=soup.find("div",{"class":"c-map__elements"})
    e1 = e.find_all("li", {"class": "m-map-item"})
    r=1
    for ee in e1:
        r=r+1
        add = ee.find("div",{"class":"m-map-item__adress"})

        #NAME
        nam = add.find("strong")
        name=str(nam.contents)
        name=name.replace("['","")
        name=name.replace("']","")
        print(name,"NAME")
        #website
        web=add.find("a")
        try:

            website=str(web.get("href"))

        except:
            website="None"
        print(website,"Website")
        #Address
        addr=str(add).replace('<div class="m-map-item__adress">',"")
        addr=addr.replace(str(web),"")
        addr=addr.replace("</div>","")
        addr = addr.replace(str(nam), "")
        addres = addr.split("<br/>")
        count=0
        for i in addres:
            count=count+1
            if count==2:
                address = addres[1].strip()
            elif count==3:
                plz=addres[2].strip()
                plz=plz.replace("                               ","")
                print(plz)
            #Phone

            cond = ee.find_all("div",{"class":"m-map-item__contact"})

        for c in cond:
            ll =c.find("a", {"class": "a-link a-link--normalcase"})
            l = ll.get("href")
            if "tel" in l:
                phone=l.replace("tel:","")
            elif "mailto" in l:
                email = l.replace("mailto:","")
        print(email,"Email","\n",phone,"Phone")
            #Email
        sh1.cell(row=r, column=1).value = name
        sh1.cell(row=r, column=2).value = plz
        sh1.cell(row=r, column=3).value = address
        sh1.cell(row=r, column=4).value = phone
        sh1.cell(row=r, column=5).value = email
        sh1.cell(row=r, column=6).value = website
        wb.save("C:\\Users\\Hp\\Desktop\\New folder (3)\\last   .xlsx")
    e2.clear()
    time.sleep(15)

